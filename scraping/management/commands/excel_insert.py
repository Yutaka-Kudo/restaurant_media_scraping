from django.core.management.base import BaseCommand
import openpyxl

from django_pandas.io import read_frame
import pandas as pd
import os
from django.http import HttpResponse
from scraping import models
from devtools import debug
import datetime
from dateutil.relativedelta import relativedelta

from scraping.views import select_dbmodel, create_df, trans_date
from scraping.models import Fes_GMB, Grg_GMB, Toro_GMB, Wanaichi_GMB, Wananakame_GMB


store_list = ["fes", "garage", "tourou", "wanaichi", "wananakame"]
# store_list = ["garage", "tourou", "wanaichi"]
# store_list = ["fes", "wananakame"]

media_list = ["gn", "hp", "tb"]
# media_list = ["hp", "tb"]
# media_list = ["gn"]

wbname = '/Users/yutakakudo/OneDrive/ユビレジチェックと媒体レポート.xlsx'
wb = openpyxl.load_workbook(wbname)
#  表の一番最後のデータの日にちを求める。
st = wb['FESレポート']  # どこのでも良い
last_date_row = [i for i in range(5, st.max_row) if st.cell(i, 1).value][-1]
last_date_span: str = st.cell(last_date_row, 1).value  # →"2021-03-07_03-13"
last_day: str = last_date_span[:5] + last_date_span[-5:]  # →"2021-03-13"
next_day: datetime.date = datetime.datetime.strptime(last_day, "%Y-%m-%d").date()+relativedelta(days=1)
day_after6: datetime.date = next_day+relativedelta(days=6)


def excel_insert_weekly(store: str, media: str):

    dbmodel, _, _ = select_dbmodel(store, media)
    # to_day: datetime.date = dbmodel.objects.latest('date').date
    qs = dbmodel.objects.filter(date__gte=next_day, date__lte=day_after6).order_by('-date')
    for i in qs:
        if i.week == "土":  # 土曜を探して起点にする
            saturday = i.date
            break

    # 過去12週間分
    # before12weeks: datetime.date = saturday - relativedelta(days=83)
    before6days: datetime.date = saturday - relativedelta(days=6)
    # _df = create_df(None, store, media, interval="weekly", start=saturday, end=before12weeks)
    _df = create_df(None, store, media, interval="weekly", start=saturday, end=before6days)

    data_list = []
    # 7日ごとの合計をリストに挿入
    for i in range(0, len(_df), 7):
        data_list.append(_df[i:i+7].sum())

    df = pd.DataFrame(data_list)
    # dates = [trans_date(s)+"_"+trans_date(s+relativedelta(days=6))[5:] for s in _df["日付"][::7]][6:]  # 6個分まで
    dates = [trans_date(s)+"_"+trans_date(s+relativedelta(days=6))[5:] for s in _df["日付"][::7]]

    if store == "fes":
        name = "FES"
    elif store == "garage":
        name = "Garage"
    elif store == "tourou":
        name = "灯篭"
    elif store == "wanaichi":
        name = "罠一目"
    elif store == "wananakame":
        name = "罠中目黒"

    st = wb[f'{name}レポート']

    if media == "gn":
        # total_pv = list(df["合計PV_SP"]+df["合計PV_PC"]+df["合計PV_app"])[6:]
        total_pv = list(df["合計PV_SP"]+df["合計PV_PC"]+df["合計PV_app"])
        data5 = list(df["店舗トップPV_SP"]+df["店舗トップPV_PC"]+df["店舗トップPV_app"])
        data21 = list(df["予約_合計件数"])
        # cvr = [(i / i2) for i, i2 in zip(data21, data5)][6:]
        cvr = [(i / i2) for i, i2 in zip(data21, data5)]

        # for i in range(6):
        #     st.cell(5+i, 1).value = dates[i]
        #     st.cell(5+i, 2).value = total_pv[i]
        #     st.cell(5+i, 4).value = cvr[i]

        st.cell(last_date_row+1, 1).value = dates[0]
        st.cell(last_date_row+1, 2).value = total_pv[0]
        st.cell(last_date_row+1, 4).value = cvr[0]

    elif media == "hp":
        # total_pv = list(df["店舗総PV_sp"]+df["店舗総PV_pc"])[6:]
        total_pv = list(df["店舗総PV_sp"]+df["店舗総PV_pc"])
        data4 = list(df["店舗TOP PV_sp"]+df["店舗TOP PV_pc"])
        data13 = list(df["予約件数_sp"]+df["予約件数_pc"])

        # cvr = [(i / i2) for i, i2 in zip(data13, data4)][6:]
        cvr = [(i / i2) for i, i2 in zip(data13, data4)]

        # for i in range(6):
        #     st.cell(5+i, 5).value = total_pv[i]
        #     st.cell(5+i, 7).value = cvr[i]

        st.cell(last_date_row+1, 5).value = total_pv[0]
        st.cell(last_date_row+1, 7).value = cvr[0]

    elif media == "tb":
        # total_pv = list(df["合計PV_sp"]+df["合計PV_pc"])[6:]
        total_pv = list(df["合計PV_sp"]+df["合計PV_pc"])
        data4 = list(df["店舗トップPV_sp"]+df["店舗トップPV_pc"])
        data13 = list(df["ネット予約"])

        # cvr = [(i / i2) for i, i2 in zip(data13, data4)][6:]
        cvr = [(i / i2) for i, i2 in zip(data13, data4)]

        # for i in range(6):
        #     st.cell(i+1, 8).value = total_pv[i]
        #     st.cell(i+1, 10).value = cvr[i]

        st.cell(last_date_row+1, 8).value = total_pv[0]
        st.cell(last_date_row+1, 10).value = cvr[0]

    wb.save(wbname)

    print(dates)
    print(total_pv)
    print(cvr)
    print(len(dates))
    print(len(total_pv))
    print(len(cvr))


def excel_insert_GMB(store: str):

    # span_id = trans_date("2021-01-31")  # これ確認
    span_id = next_day  # これ確認
    debug(span_id)

    if store == "fes":
        dbmodel = Fes_GMB
        name = "FES"
    elif store == "garage":
        dbmodel = Grg_GMB
        name = "Garage"
    elif store == "tourou":
        dbmodel = Toro_GMB
        name = "灯篭"
    elif store == "wanaichi":
        dbmodel = Wanaichi_GMB
        name = "罠一目"
    elif store == "wananakame":
        dbmodel = Wananakame_GMB
        name = "罠中目黒"
    st = wb[f'{name}レポート']
    qs = dbmodel.objects.filter(span_id=span_id)

    st.cell(last_date_row+1, 11).value = qs[0].total_show
    span_id += relativedelta(days=7)
    qs = dbmodel.objects.filter(span_id=span_id)

    wb.save(wbname)

    # # 最初にやったやつ。まとめて登録
    # span_id = trans_date("2021-01-31")  # これ確認

    # if store == "fes":
    #     dbmodel = Fes_GMB
    #     name = "FES"
    # elif store == "garage":
    #     dbmodel = Grg_GMB
    #     name = "Garage"
    # elif store == "tourou":
    #     dbmodel = Toro_GMB
    #     name = "灯篭"
    # elif store == "wanaichi":
    #     dbmodel = Wanaichi_GMB
    #     name = "罠一目"
    # elif store == "wananakame":
    #     dbmodel = Wananakame_GMB
    #     name = "罠中目黒"
    # st = wb[f'{name}レポート']
    # start_row = 5
    # qs = dbmodel.objects.filter(span_id=span_id)

    # while qs:  # データなくなるまで
    #     st.cell(start_row, 11).value = qs[0].total_show
    #     start_row += 1
    #     span_id += relativedelta(days=7)
    #     qs = dbmodel.objects.filter(span_id=span_id)

    # wb.save(wbname)


class Command(BaseCommand):  # コマンド python manage.py get_insert_db
    def handle(self, *args, **options):

        debug(last_date_row, next_day)

        for s in store_list:
            for m in media_list:
                debug(s, m)
                excel_insert_weekly(s, m)

        for s in store_list:  # GMB用
            excel_insert_GMB(s)
